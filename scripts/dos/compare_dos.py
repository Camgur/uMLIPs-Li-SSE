import os, glob, argparse
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from sumo.io.castep import read_dos

import sys
sys.path.insert(0, str(Path(__file__).parents[2]))
from utils.config import (
    GAUSSIAN_SIGMA,
    DEFAULT_STRUCTURE, MODELS, COLORS
)

# Helpers
def find_bands_file(directory):
    """Find .bands file in directory."""
    files = glob.glob(os.path.join(directory, "*.bands"))
    return files[0] if files else None

def load_and_total(bands_file, gaussian, emin, emax):
    """Load CASTEP DOS with SUMO and return dos_obj for native VBM/CBM access."""
    dos_obj, _ = read_dos(bands_file, gaussian=gaussian, emin=emin, emax=emax, total_only=False)
    energies = np.array(dos_obj.energies)
    # sum spins if dict
    if isinstance(dos_obj.densities, dict):
        total = sum(np.array(arr) for arr in dos_obj.densities.values())
    else:
        total = np.array(dos_obj.densities)
        if total.ndim > 1:
            total = total.sum(axis=0)
    return energies, np.array(total), float(dos_obj.efermi), dos_obj

def compare_and_plot(structure, models, output_dir="results/dos"):
    """Plot DOS for ML models and DFT, export figure and Excel."""
    os.makedirs(output_dir, exist_ok=True)
    
    fig, ax = plt.subplots(figsize=(10, 6))
    
    wb = Workbook()
    ws = wb.active
    ws.title = structure
    ws['A1'] = 'Model'
    ws['B1'] = 'VBM (eV)'
    ws['C1'] = 'CBM (eV)'
    ws['D1'] = 'Bandgap (eV)'
    
    row = 2
    
    # Plot ML models first
    for model in models:
        model_dir = f"MACE/elastic/{structure}/{model}"
        if not os.path.isdir(model_dir):
            print(f"[WARNING] No directory for {structure} / {model}")
            continue
        
        bands_file = find_bands_file(model_dir)
        if not bands_file:
            print(f"[WARNING] No bands file for {structure} / {model}")
            continue
        
        try:
            energies, total_dos, efermi, dos_obj = load_and_total(bands_file, GAUSSIAN_SIGMA, -15, 15)
            e_rel = energies - efermi
            
            # Use SUMO's native VBM/CBM detection
            vbm = float(dos_obj.vbm) if dos_obj.vbm is not None else 0.0
            cbm = float(dos_obj.cbm) if dos_obj.cbm is not None else np.nan
            bandgap = cbm - vbm if not np.isnan(cbm) else np.nan
            
            # Plot
            ax.plot(e_rel, total_dos, 
                   color=COLORS.get(model, '#000000'),
                   label=f"{model}: {bandgap:.2f} eV" if not np.isnan(bandgap) else model,
                   linewidth=2)
            
            # Write to Excel
            ws[f'A{row}'] = model
            ws[f'B{row}'] = round(vbm, 4)
            ws[f'C{row}'] = round(cbm, 4) if not np.isnan(cbm) else None
            ws[f'D{row}'] = round(bandgap, 4) if not np.isnan(bandgap) else None
            row += 1
            
            print(f"Plotted {structure} / {model}: VBM={vbm:.3f}, CBM={cbm:.3f}, Eg={bandgap:.3f} eV")
            
        except Exception as exc:
            print(f"[ERROR] Could not process {structure} / {model}: {exc}")
    
    # Plot DFT last
    dft_dir = f"castep/{structure}"
    if os.path.isdir(dft_dir):
        bands_file = find_bands_file(dft_dir)
        if bands_file:
            try:
                energies, total_dos, efermi, dos_obj = load_and_total(bands_file, GAUSSIAN_SIGMA, -15, 15)
                e_rel = energies - efermi
                
                # Use SUMO's native VBM/CBM detection
                vbm = float(dos_obj.vbm) if dos_obj.vbm is not None else 0.0
                cbm = float(dos_obj.cbm) if dos_obj.cbm is not None else np.nan
                bandgap = cbm - vbm if not np.isnan(cbm) else np.nan
                
                # Plot DFT on top
                ax.plot(e_rel, total_dos, 
                       color=COLORS['dft'],
                       label=f"DFT: {bandgap:.2f} eV" if not np.isnan(bandgap) else "DFT",
                       linewidth=2.5,
                       linestyle='solid')
                
                # Write to Excel
                ws[f'A{row}'] = 'DFT'
                ws[f'B{row}'] = round(vbm, 4)
                ws[f'C{row}'] = round(cbm, 4) if not np.isnan(cbm) else None
                ws[f'D{row}'] = round(bandgap, 4) if not np.isnan(bandgap) else None
                
                print(f"Plotted {structure} / DFT: VBM={vbm:.3f}, CBM={cbm:.3f}, Eg={bandgap:.3f} eV")
                
            except Exception as exc:
                print(f"[ERROR] Could not process {structure} / DFT: {exc}")
    
    # Format plot
    ax.set_xlabel("Energy (eV)", fontsize=14)
    ax.set_ylabel("DOS (states/eV)", fontsize=14)
    ax.set_title(f"Density of States: {structure}", fontsize=16)
    ax.set_xlim(-15, 15)
    ax.set_ylim(bottom=0)
    ax.legend(fontsize=11, loc='best')
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    
    # Save figure
    fig_path = os.path.join(output_dir, f"dos_{structure}.png")
    fig.savefig(fig_path, dpi=300)
    print(f"Saved figure to {fig_path}")
    plt.close()
    
    # Save Excel
    excel_path = os.path.join(output_dir, f"bandgaps_{structure}.xlsx")
    wb.save(excel_path)
    print(f"Saved Excel to {excel_path}")


def main():
    parser = argparse.ArgumentParser(description="Plot DOS comparison for a structure")
    parser.add_argument("--structure", default=DEFAULT_STRUCTURE, 
                       help=f"Structure name (default: {DEFAULT_STRUCTURE})")
    parser.add_argument("--output-dir", default="results/dos",
                       help="Output directory for figures and data")
    args = parser.parse_args()
    
    compare_and_plot(args.structure, MODELS, args.output_dir)


if __name__ == "__main__":
    main()
